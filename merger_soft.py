import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import gspread

def insert_data(filename):
    try:
        data = []
        with open(filename, 'r') as f:
            reader = csv.reader(f)
            next(reader,None)  # Skip the header row
            for row in reader:
                data.append(row)

        if len(data) > 0:
            gh = gspread.service_account("./credentials/credentials_file.json")
            wk = gh.open("Teddy Scraper")
            sh = wk.worksheet("Deal software")
            sh.insert_rows(data, 2)
    except Exception as e:
        print(f"An error occurred: {e}")

def insert_csv_to_excel(csv_filename, excel_filename):
    # Function to insert rows
    def insert_rows(ws, dataframe, row_start=1):
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=False), row_start):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Read the CSV file
    csv_data = pd.read_csv(csv_filename)

    # Load the Excel file
    book = load_workbook(excel_filename)
    sheet = book.active

    # Insert CSV data into Excel
    insert_rows(sheet, csv_data, row_start=2)  # Start from row 2 to keep the header

    # Save the updated Excel file
    updated_excel_file = excel_filename
    book.save(updated_excel_file)
    print(f"Updated Excel file saved as: {updated_excel_file}")

if __name__ == "__main__":
    insert_csv_to_excel('./soft_data.csv','./Deal_softwares.xlsx')
    insert_data('./soft_data.csv')